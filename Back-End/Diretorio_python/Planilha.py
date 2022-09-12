import openpyxl

book = openpyxl.Workbook()

book.create_sheet('Sheet1')

print(book.sheetnames)

funcionarios = book['Sheet1']

funcionarios.append(['EMPRESA', 'R.E', 'NOME', 'STATUS', 'DOC P'])

documentos = ("Ficha de registro", "Contrato de trabalho", "Docs Diversos", "Sindicato Carta", "VT Opção", "EPI's", "ASO", "LGPD")
dados = []
funcionario = []

# print('Digite os dados: ')
# for a in range(0, 5):
#     dado = input('Digite os dados -- > ')

def dados_deum_fucionario():
    print('Digite os dados: ')
    empresa = input('Empresa: \n-- > ')
    dados.append(empresa)
    re = input('R.E: \n-- > ')
    dados.append(re)
    nome = input('Nome: \n-- > ')
    dados.append(nome)
    status = input('Status: \n-- > ')
    dados.append(status)

    for b in range(0, 8):
        dados.append(documentos[b])
        funcionario = dados
        dados.pop

    print(dados)
    print(funcionario)

# for b in range(0, 8):
#     dados.append(documentos[a])
#     funcionario.append(dados)
#     dados.pop()
#funcionarios.append(funcionario)

dados_deum_fucionario()

