#diretorio da minha máquina


import openpyxl

book1 = openpyxl.Workbook()

book2 = openpyxl.load_workbook('dados.xlsx')
#print(book2.sheetnames)

funcionarios = book1['Sheet']

tabelinha = book2['Planilha1']
print

linhas = []
                #pegando os dados de linhas, atribuindo as linhas da tabela à lista "linhas"
for rows in tabelinha.iter_rows(min_row=2):
    linha = []              #executado a cada linha
    for cell in rows:       #executado a cada celula
        linha.append(cell.value)
    linhas.append(linha)

print(linhas)
def tratar_funcionario():
    funcionarios.append(["EMPRESA", "R.E", "NOME", "STATUS", "DOC P"])
    documentos = ["Ficha de registro", "Contrato de trabalho", "Docs Diversos", "Sindicato Carta", "VT Opção", "ASO", "LGPD"]
    dados = []
    lista = []
    funcionario = []

    print('Digite os dados: ')

    empresa = str(input("Empresa: \n-- > "))
    dados.append(empresa)

    re = str(input("R.E: \n-- > "))
    dados.append(re)

    nome = str(input("Nome: \n-- > "))
    dados.append(nome)

    status = str(input("Status: \n-- > "))
    dados.append(status)

    for a in range(len(documentos)):
        lista = dados.copy()
        lista.append(documentos[a])
        funcionarios.append(lista)

book1.save('Tabela_de_teste_1.xlsx')